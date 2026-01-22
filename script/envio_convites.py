import requests
from openpyxl import load_workbook
import urllib.parse
import time
import os
from dotenv import load_dotenv

# ----------------------------------------
# Carregar variáveis do arquivo .env
# ----------------------------------------
load_dotenv()
API_KEY = os.getenv("API_KEY")
if not API_KEY:
    raise ValueError("API_KEY não encontrada no arquivo .env")

# ----------------------------------------
# Configurações da API
# ----------------------------------------
API_URL = "http://localhost:8080"
INSTANCE = "convites"

HEADERS = {
    "apikey": API_KEY,
    "Content-Type": "application/json"
}

# ----------------------------------------
# Caminho relativo para o arquivo Excel
# ----------------------------------------
ARQUIVO = "planilhas\convidados.xlsx"
ABA = "lista"

# Carregar planilha
wb = load_workbook(ARQUIVO)
ws = wb[ABA]

# ----------------------------------------
# Função para transformar links do Google Drive em download direto
# ----------------------------------------
def drive_download(link):
    if "drive.google.com" in link and "/file/d/" in link:
        file_id = link.split("/file/d/")[1].split("/")[0]
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    return link

# ----------------------------------------
# Criar dicionário de grupos únicos
# ----------------------------------------
grupos = {}
for row in ws.iter_rows(min_row=2):
    grupo = row[0].value
    nome = row[1].value
    telefone = str(row[4].value)
    Responsavel = row[3].value
    cell_pdf = row[7]

    if not telefone:
        continue

    # Link dinâmico do formulário
    BASE_FORM_URL = "https://docs.google.com/forms/d/e/EXEMPLO_DO_FORMULARIO/viewform?usp=pp_url"
    link_form = f"{BASE_FORM_URL}&entry.47010176={urllib.parse.quote(str(grupo))}&entry.163752669={urllib.parse.quote(str(nome))}"

    # Link do PDF
    link_pdf = ""
    if cell_pdf.hyperlink:
        link_pdf = drive_download(cell_pdf.hyperlink.target)

    # Agrupar por grupo
    if grupo not in grupos:
        grupos[grupo] = {
            "telefone": telefone,
            "nomes": [nome],
            "link_form": link_form,
            "link_pdf": link_pdf
        }
    else:
        grupos[grupo]["nomes"].append(nome)

total = len(grupos)

# ----------------------------------------
# Enviar convites
# ----------------------------------------
for i, (grupo, info) in enumerate(grupos.items(), 1):
    telefone = info["telefone"]
    nomes = info["nomes"]
    link_form = info["link_form"]
    link_pdf = info["link_pdf"]

    lista_convidados = "\n".join(f"🎟️ {n}" for n in nomes)

    if "& FAMILIA" in grupo.upper():
        saudacao = grupo
    else:
        saudacao = f"Olá {nomes[0]}, da família {grupo}!"

    texto = f"""💌 Convite Especial! Confirme sua presença 🎉👰🤵

{saudacao} 💌

* Texto do convite *

👉 Link para Ônibus: {link_form}

🎟️ Vale-Convite  
Este convite é válido exclusivamente para as pessoas listadas abaixo:

{lista_convidados}

* Agradecimentos e informações adicionais *
"""

    # Enviar TEXTO
    requests.post(
        f"{API_URL}/message/sendText/{INSTANCE}",
        headers=HEADERS,
        json={"number": telefone, "textMessage": {"text": texto}}
    )

    # Enviar PDF
    if link_pdf:
        requests.post(
            f"{API_URL}/message/sendMedia/{INSTANCE}",
            headers=HEADERS,
            json={
                "number": telefone,
                "mediaMessage": {
                    "mediatype": "document",
                    "fileName": "Convite_Casamento.pdf",
                    "media": link_pdf
                }
            }
        )

    # Print com contador
    print(f"[{i}/{total}] ✅ Convite enviado para {grupo}")

    # Delay de 2 segundos entre envios
    time.sleep(2)

print("🎉 TODOS OS CONVITES ENVIADOS")
